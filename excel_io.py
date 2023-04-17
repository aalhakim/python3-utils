#!/usr/bin/env python3
"""
Example code on how to read from and write to a Microsoft Excel document.
"""

# Standard library imports
import os

# Third-party libray imports
import openpyxl


###############################################################################
def open_xlsx(filepath):
    """Open an XLSX file
    This function uses `openppyxl` which is only compatible with XLSX
    files.
    PARAMETERS
    ==========
    filepath: <str>
        The absolute path of the file you want to open.
    RETURNS
    =======
    This function will return an <openpyxl.workbook.Workbook> object.
    """

    f_extension = filepath.split(".")[-1].lower()
    assert (
        f_extension == "xlsx"
    ), "ERROR: File '{}' has extension `{}` but should be `xlsx`.".format(
        filepath, f_extension
    )
    return openpyxl.load_workbook(filepath)


def create_xlsx(filepath):
    """Create a new blank XLSX file
    PARAMETERS
    ==========
    filepath: <str>
        The location when the new XLSX file will be saved.
    RETURNS
    =======
    This function will return the new blank <openpyxl.workbook.Workbook>
    object.
    """

    # Create a save the new workbook
    workbook = openpyxl.Workbook()
    save_workbook(workbook, filepath)
    return workbook


def save_workbook(workbook, filepath):
    """Save a <openpyxl.workbook.Workbook> object to as an XLSX file
    PARAMETERS
    ==========
    filepath: <str>
        The location when the new XLSX file will be saved.
    """

    # Force the correct file extension
    filename = os.path.basename(filepath)
    f_split = filename.split(".")
    f_extension = f_split[-1]
    if f_split == 1:
        filepath = filepath + ".xlsx"

    elif f_extension.lower() != "xlsx":
        filepath = filepath.strip(f_extension) + "xlsx"

    else:
        pass  # No action

    workbook.save(filepath)


def get_worksheet_names(workbook):
    """Return a list of sheet names in an XLXS workbook
    PARAMETERS
    ==========
    workbook: <openpyxl.workbook.Workbook>
        An XLSX workbook object
    RETURNS
    =======
    This function will return a [list of <str>] of all the worksheet
    names in the XLSX workbook.
    """

    return workbook.sheetnames


def get_worksheet(workbook, worksheet_name):
    """Retrieve data from the given worksheet.
    PARAMETERS
    ==========
    workbook: <openpyxl.workbook.Workbook>
        An XLSX workbook object
    worksheet_name: <str>
        The name of a valid <openpyxl.worksheet.Worksheet> object.
    RETURNS
    =======
    This function will return a [] of all data in the worksheet.
    """
    return workbook[worksheet_name]


def get_row_data(worksheet, row=1, length=50):
    """Return a list of data from row number `row`
    PARAMETERS
    ==========
    worksheet: <openpyxl.worksheet.Worksheet>
        A Worksheet object.
    row: <int>
        ID of the row to retrieve data for
    length: <int>
        The number of columns to retieve data for
    RETURNS
    =======
    This function will return a [list of <str>] of length `length`
    containing data from the cells of row `row`.
    """

    data = []
    for col in worksheet.iter_cols(min_row=row, max_col=length, max_row=row):
        for cell in col:
            cell_value = cell.value
            if cell_value is None:
                data.append("")
            else:
                data.append(cell_value)
    return data


def get_header_names(worksheet):
    """Return a list of header names from an <openpyxl.worksheet.Worksheet>
    object
    The end of the header group is determined by the first Null cell
    in the first row of the document. i.e. the first time a value
    is None is find will determine the end of the headers.
    PARAMETERS
    ==========
    worksheet: <openpyxl.worksheet.Worksheet>
        A worksheet object.
    RETURNS
    =======
    This function will return a [list of <str>] with a length
    dependant on the number of headers in the worksheet table,
    although limited to 50 items.
    """

    headers = []
    for col in worksheet.iter_cols(min_row=1, max_col=50, max_row=1):
        for cell in col:
            cell_value = cell.value
            if cell_value is None:
                break
            else:
                headers.append(cell.value)
    return headers


def get_max_columns(worksheet, row=1):
    """Identify the number of columns in a worksheet table
    PARAMETERS
    ==========
    worksheet: <openpyxl.worksheet.Worksheet>
        A worksheet object.
    row: <int>
        Select which row to use to count the table length.
    RETURNS
    =======
    This function will return an <int> value to indicate the number
    of columns in given row.
    """

    length = 0
    null_count = 0
    for col in worksheet.iter_cols(min_row=row, max_row=row):
        for cell in col:
            # Stop counting if there are 10 consecutive Null cells
            if cell.value is None:
                null_count += 1
                if null_count == 10:
                    break
            else:
                length += null_count + 1
                null_count = 0
    return length


def get_max_rows(worksheet, column=1):
    """Identify the number of rows in a worksheet table
    PARAMETERS
    ==========
    worksheet: <openpyxl.worksheet.Worksheet>
        A worksheet object.
    column: <int>
        Select which column to use to count the table height.
    RETURNS
    =======
    This function will return an <int> value to indicate the number
    of rows in the given column.
    """

    height = 0
    null_count = 0
    for row in worksheet.iter_rows(min_col=column, max_col=column):
        for cell in row:
            # Stop counting if there are 10 consecutive Null cells
            if cell.value is None:
                null_count += 1
                if null_count == 10:
                    break
            else:
                height += null_count + 1
                null_count = 0
    return height


########################################################################
if __name__ == "__main__":
    TEST_FILE = "transistors.xlsx"
    workbook = open_xlsx(TEST_FILE)
    worksheet_names = get_worksheet_names(workbook)
    print("Worksheets: {}".format(worksheet_names))

    for ws_name in worksheet_names:
        worksheet = workbook[ws_name]

        table_length = get_max_columns(worksheet)
        table_height = get_max_rows(worksheet)
        print("Table Size: Rows={}, Cols={}".format(table_height, table_length))

        headers = get_header_names(worksheet)
        print("Headers: ", end=" ")
        for header in headers:
            print("'{}'".format(header), end=" ")
        print()

        for row in range(2, table_height + 1):
            print()
            print(row, get_row_data(worksheet, row, table_length))

    NEW_FILE = "./test.xlsx"
    wb2 = create_xlsx(NEW_FILE)
    print(get_worksheet_names(wb2))
